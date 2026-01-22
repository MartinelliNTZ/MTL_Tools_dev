import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Data for all 11 classes with their methods
classes_data = {
    "VectorLayerSource": {
        "type": "vector",
        "scope": "Entrada, saída e origem de camadas vetoriais",
        "methods": [
            ("load_vector_layer_from_file", "Entrada", "Arquivo", "Camada Vetorial"),
            ("load_vector_layer_from_database", "Entrada", "Banco de Dados", "Camada Vetorial"),
            ("create_memory_layer", "Saída", "Memória", "Camada Vetorial"),
            ("save_vector_layer_to_file", "Saída", "Camada Vetorial", "Arquivo"),
            ("save_vector_layer_to_database", "Saída", "Camada Vetorial", "Banco de Dados"),
            ("clone_vector_layer", "Transformação", "Camada Vetorial", "Camada Vetorial"),
            ("validate_layer_integrity", "Validação", "Camada Vetorial", "Booleano"),
            ("get_layer_source_uri", "Leitura", "Camada Vetorial", "String (URI)"),
            ("check_file_format_compatibility", "Validação", "Arquivo/Formato", "Booleano"),
            ("reload_layer_from_source", "Transformação", "Camada Vetorial", "Camada Vetorial"),
            ("export_layer_statistics", "Saída", "Camada Vetorial", "Arquivo Relatório"),
            ("validate_geometry_before_save", "Validação", "Camada Vetorial", "Booleano"),
        ]
    },
    "VectorLayerProjection": {
        "type": "vector",
        "scope": "CRS, reprojeção e unidades",
        "methods": [
            ("get_layer_crs", "Leitura", "Camada Vetorial", "CRS"),
            ("set_layer_crs", "Transformação", "Camada Vetorial/CRS", "Camada Vetorial"),
            ("reproject_layer_to_crs", "Transformação", "Camada Vetorial/CRS", "Camada Vetorial"),
            ("validate_crs_compatibility", "Validação", "Camada Vetorial x2", "Booleano"),
            ("get_layer_unit_type", "Leitura", "Camada Vetorial", "String (Unidade)"),
            ("convert_distance_unit", "Transformação", "Valor/Unidades", "Valor Convertido"),
            ("convert_area_unit", "Transformação", "Valor/Unidades", "Valor Convertido"),
            ("get_layer_extent_in_different_crs", "Cálculo", "Camada Vetorial/CRS", "Extent"),
            ("check_layer_is_geographic_crs", "Validação", "Camada Vetorial", "Booleano"),
            ("check_layer_is_projected_crs", "Validação", "Camada Vetorial", "Booleano"),
            ("get_crs_axis_order", "Leitura", "CRS", "String (XY/YX)"),
            ("list_common_crs_for_region", "Leitura", "Código Região", "Lista CRS"),
        ]
    },
    "VectorLayerGeometry": {
        "type": "vector",
        "scope": "Transformações geométricas",
        "methods": [
            ("create_buffer_geometry", "Transformação", "Camada Vetorial/Distância", "Camada Vetorial"),
            ("dissolve_geometries_by_attribute", "Transformação", "Camada Vetorial/Campo", "Camada Vetorial"),
            ("merge_geometries", "Transformação", "Geometrias x2+", "Geometria"),
            ("explode_multipart_features", "Transformação", "Camada Vetorial", "Camada Vetorial"),
            ("simplify_geometry", "Transformação", "Camada Vetorial/Tolerância", "Camada Vetorial"),
            ("smooth_geometry", "Transformação", "Camada Vetorial/Iterações", "Camada Vetorial"),
            ("validate_geometry", "Validação", "Geometria", "Booleano"),
            ("fix_invalid_geometry", "Transformação", "Geometria Inválida", "Geometria Corrigida"),
            ("get_geometry_intersection", "Cálculo", "Geometria x2", "Geometria"),
            ("get_geometry_union", "Cálculo", "Geometria x2", "Geometria"),
            ("get_geometry_difference", "Cálculo", "Geometria x2", "Geometria"),
            ("convert_geometry_type", "Transformação", "Camada Vetorial/Tipo", "Camada Vetorial"),
        ]
    },
    "VectorLayerMetrics": {
        "type": "vector",
        "scope": "Leitura e cálculos espaciais",
        "methods": [
            ("calculate_feature_area", "Cálculo", "Feature", "Float (Área)"),
            ("calculate_feature_length", "Cálculo", "Feature", "Float (Comprimento)"),
            ("calculate_feature_perimeter", "Cálculo", "Feature", "Float (Perímetro)"),
            ("get_layer_total_area", "Cálculo", "Camada Vetorial", "Float (Área Total)"),
            ("get_layer_total_length", "Cálculo", "Camada Vetorial", "Float (Comprimento Total)"),
            ("get_layer_extent_area", "Cálculo", "Camada Vetorial", "Float (Área Extent)"),
            ("calculate_feature_count_by_area", "Análise", "Camada Vetorial/Ranges", "Dicionário"),
            ("get_feature_centroid", "Cálculo", "Feature", "Ponto"),
            ("get_layer_centroid", "Cálculo", "Camada Vetorial", "Ponto"),
            ("calculate_distance_between_features", "Cálculo", "Feature x2", "Float (Distância)"),
            ("get_layer_density_statistics", "Análise", "Camada Vetorial/Tamanho Grade", "Estatísticas"),
            ("analyze_spatial_distribution", "Análise", "Camada Vetorial", "Estatísticas"),
        ]
    },
    "VectorLayerAttributes": {
        "type": "vector",
        "scope": "Campos e atributos",
        "methods": [
            ("get_layer_fields", "Leitura", "Camada Vetorial", "Lista Campos"),
            ("create_new_field", "Transformação", "Camada Vetorial/Campo", "Booleano"),
            ("delete_field", "Transformação", "Camada Vetorial/Campo", "Booleano"),
            ("rename_field", "Transformação", "Camada Vetorial/Campo x2", "Booleano"),
            ("get_field_values", "Leitura", "Camada Vetorial/Campo", "Lista Valores"),
            ("calculate_field_statistics", "Cálculo", "Camada Vetorial/Campo", "Estatísticas"),
            ("validate_field_data_type", "Validação", "Camada Vetorial/Campo/Tipo", "Booleano"),
            ("update_field_values", "Transformação", "Camada Vetorial/Campo/Valores", "Booleano"),
            ("check_for_duplicate_values", "Validação", "Camada Vetorial/Campo", "Lista"),
            ("get_feature_attribute", "Leitura", "Feature/Campo", "Valor"),
            ("set_feature_attribute", "Transformação", "Feature/Campo/Valor", "Booleano"),
            ("export_attributes_to_table", "Saída", "Camada Vetorial", "Arquivo Tabular"),
        ]
    },
    "RasterLayerSource": {
        "type": "raster",
        "scope": "Carregamento, salvamento e criação de rasters",
        "methods": [
            ("load_raster_from_file", "Entrada", "Arquivo", "Camada Raster"),
            ("load_raster_from_url", "Entrada", "URL", "Camada Raster"),
            ("create_empty_raster", "Saída", "Parâmetros Espaciais", "Camada Raster"),
            ("save_raster_to_file", "Saída", "Camada Raster", "Arquivo"),
            ("get_raster_bands_info", "Leitura", "Camada Raster", "Info Bandas"),
            ("validate_raster_file", "Validação", "Arquivo", "Booleano"),
            ("get_raster_nodata_value", "Leitura", "Camada Raster/Banda", "Valor Nodata"),
            ("set_raster_nodata_value", "Transformação", "Camada Raster/Banda/Valor", "Booleano"),
            ("copy_raster_structure", "Transformação", "Camada Raster", "Arquivo/Camada Raster"),
            ("clone_raster", "Transformação", "Camada Raster", "Camada Raster"),
            ("get_raster_file_size", "Leitura", "Arquivo", "Int (Bytes)"),
            ("export_raster_metadata", "Saída", "Camada Raster", "Arquivo Documentação"),
        ]
    },
    "RasterLayerProjection": {
        "type": "raster",
        "scope": "CRS, resolução, alinhamento e extent",
        "methods": [
            ("get_raster_crs", "Leitura", "Camada Raster", "CRS"),
            ("set_raster_crs", "Transformação", "Camada Raster/CRS", "Camada Raster"),
            ("reproject_raster_to_crs", "Transformação", "Camada Raster/CRS/Método", "Camada Raster"),
            ("get_raster_resolution", "Leitura", "Camada Raster", "Tupla (X, Y)"),
            ("set_raster_resolution", "Transformação", "Camada Raster/Resoluções", "Camada Raster"),
            ("get_raster_extent", "Leitura", "Camada Raster", "Extent"),
            ("align_raster_to_grid", "Transformação", "Camada Raster/Parâmetros Grade", "Camada Raster"),
            ("align_raster_to_reference", "Transformação", "Camada Raster x2", "Camada Raster"),
            ("get_raster_geotransform", "Leitura", "Camada Raster", "Geotransform"),
            ("set_raster_geotransform", "Transformação", "Camada Raster/Geotransform", "Booleano"),
            ("validate_raster_alignment", "Validação", "Camada Raster x2", "Booleano"),
            ("convert_pixel_to_geographic_coordinates", "Transformação", "Camada Raster/Pixel", "Coordenadas"),
        ]
    },
    "RasterLayerProcessing": {
        "type": "raster",
        "scope": "Processamento raster destrutivo",
        "methods": [
            ("apply_raster_algebra", "Transformação", "Camada Raster x2/Operação", "Camada Raster"),
            ("reclassify_raster_values", "Transformação", "Camada Raster/Regras", "Camada Raster"),
            ("apply_raster_mask", "Transformação", "Camada Raster x2", "Camada Raster"),
            ("apply_band_math_expression", "Transformação", "Camada Raster/Expressão", "Camada Raster"),
            ("combine_rasters", "Transformação", "Camada Raster x2+/Método", "Camada Raster"),
            ("apply_focal_filter", "Transformação", "Camada Raster/Parâmetros", "Camada Raster"),
            ("normalize_raster_values", "Transformação", "Camada Raster/Intervalo", "Camada Raster"),
            ("stretch_raster_histogram", "Transformação", "Camada Raster/Percentis", "Camada Raster"),
            ("apply_threshold_classification", "Transformação", "Camada Raster/Thresholds", "Camada Raster"),
            ("invert_raster_values", "Transformação", "Camada Raster", "Camada Raster"),
            ("fill_raster_nodata", "Transformação", "Camada Raster/Método", "Camada Raster"),
            ("apply_raster_reclass_table", "Transformação", "Camada Raster/Arquivo Tabela", "Camada Raster"),
        ]
    },
    "RasterLayerMetrics": {
        "type": "raster",
        "scope": "Estatísticas raster",
        "methods": [
            ("get_raster_min_max_values", "Cálculo", "Camada Raster/Banda", "Tupla (Min, Max)"),
            ("calculate_raster_mean", "Cálculo", "Camada Raster/Banda", "Float (Média)"),
            ("calculate_raster_standard_deviation", "Cálculo", "Camada Raster/Banda", "Float (Desvio Padrão)"),
            ("generate_raster_histogram", "Análise", "Camada Raster/Banda/Bins", "Histograma"),
            ("calculate_area_by_class", "Cálculo", "Camada Raster/Tamanho Pixel", "Dicionário"),
            ("get_unique_values_count", "Cálculo", "Camada Raster/Banda", "Int"),
            ("calculate_percentage_by_class", "Cálculo", "Camada Raster/Banda", "Dicionário"),
            ("analyze_raster_correlation", "Análise", "Camada Raster x2", "Float (Correlação)"),
            ("calculate_nodata_percentage", "Cálculo", "Camada Raster/Banda", "Float (Percentagem)"),
            ("get_raster_median_value", "Cálculo", "Camada Raster/Banda", "Float (Mediana)"),
            ("calculate_raster_variance", "Cálculo", "Camada Raster/Banda", "Float (Variância)"),
            ("generate_statistical_summary", "Análise", "Camada Raster/Banda", "Dicionário"),
        ]
    },
    "RasterVectorBridge": {
        "type": "hybrid",
        "scope": "Integração raster ↔ vetor",
        "methods": [
            ("rasterize_vector_layer", "Transformação", "Camada Vetorial", "Camada Raster"),
            ("polygonize_raster", "Transformação", "Camada Raster", "Camada Vetorial"),
            ("extract_zonal_statistics", "Análise", "Camada Raster + Vetorial", "Estatísticas"),
            ("clip_raster_by_vector", "Transformação", "Camada Raster + Vetorial", "Camada Raster"),
            ("sample_raster_at_points", "Transformação", "Camada Raster + Pontos", "Camada Vetorial"),
            ("calculate_raster_value_within_polygon", "Cálculo", "Camada Raster + Geometria", "Valor"),
            ("convert_raster_to_point_cloud", "Transformação", "Camada Raster", "Camada Vetorial"),
            ("validate_raster_vector_alignment", "Validação", "Camada Raster + Vetorial", "Booleano"),
            ("create_vector_from_raster_edges", "Transformação", "Camada Raster", "Camada Vetorial"),
            ("intersect_raster_with_vector_buffer", "Análise", "Camada Raster + Vetorial", "Estatísticas"),
            ("export_zonal_statistics_table", "Saída", "Estatísticas", "Arquivo Tabular"),
            ("blend_multiple_rasters_by_vector_mask", "Transformação", "Camada Raster x2+ + Vetorial", "Camada Raster"),
        ]
    },
    "RasterLayerRendering": {
        "type": "raster",
        "scope": "Simbologia e visualização raster",
        "methods": [
            ("apply_single_band_renderer", "Renderização", "Camada Raster/Banda/Rampa", "Camada Raster"),
            ("apply_multiband_renderer", "Renderização", "Camada Raster/Bandas RGB", "Camada Raster"),
            ("apply_paletted_renderer", "Renderização", "Camada Raster/Paleta", "Camada Raster"),
            ("set_raster_opacity", "Renderização", "Camada Raster/Opacidade", "Camada Raster"),
            ("set_band_opacity", "Renderização", "Camada Raster/Banda/Opacidade", "Camada Raster"),
            ("apply_color_ramp", "Renderização", "Camada Raster/Banda/Rampa", "Camada Raster"),
            ("set_raster_contrast_enhancement", "Renderização", "Camada Raster/Parâmetros", "Camada Raster"),
            ("create_color_table_from_values", "Renderização", "Valores/Cores", "Tabela Cores"),
            ("apply_hillshade_effect", "Renderização", "Camada Raster/Parâmetros", "Camada Raster"),
            ("generate_legend_for_raster", "Renderização", "Camada Raster/Formato", "Legenda"),
            ("apply_discrete_color_scheme", "Renderização", "Camada Raster/Banda/Classes", "Camada Raster"),
            ("export_rendering_style_to_file", "Saída", "Camada Raster", "Arquivo QML"),
        ]
    },
}

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Métodos Utilitários"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
vector_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
raster_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
hybrid_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Headers
headers = ["Classe", "Método", "Escopo Classe", "Tipo", "Operação", "Entrada", "Saída"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment

# Set column widths
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 25
ws.column_dimensions["G"].width = 25

# Populate data
row = 2
for class_name, class_info in sorted(classes_data.items()):
    for method_name, operation, entrada, saida in class_info["methods"]:
        ws.cell(row=row, column=1).value = class_name
        ws.cell(row=row, column=2).value = method_name
        ws.cell(row=row, column=3).value = class_info["scope"]
        ws.cell(row=row, column=4).value = class_info["type"].upper()
        ws.cell(row=row, column=5).value = operation
        ws.cell(row=row, column=6).value = entrada
        ws.cell(row=row, column=7).value = saida
        
        # Apply color based on type
        if class_info["type"] == "vector":
            fill = vector_fill
        elif class_info["type"] == "raster":
            fill = raster_fill
        else:
            fill = hybrid_fill
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).alignment = left_alignment
        
        row += 1

# Add legend at the bottom
legend_row = row + 2
ws.cell(row=legend_row, column=1).value = "Legenda:"
ws.cell(row=legend_row, column=1).font = Font(bold=True)

legend_row += 1
vector_cell = ws.cell(row=legend_row, column=1)
vector_cell.value = "VECTOR"
vector_cell.fill = vector_fill
ws.cell(row=legend_row, column=2).value = "Utilitários de camadas vetoriais"

legend_row += 1
raster_cell = ws.cell(row=legend_row, column=1)
raster_cell.value = "RASTER"
raster_cell.fill = raster_fill
ws.cell(row=legend_row, column=2).value = "Utilitários de camadas raster"

legend_row += 1
hybrid_cell = ws.cell(row=legend_row, column=1)
hybrid_cell.value = "HYBRID"
hybrid_cell.fill = hybrid_fill
ws.cell(row=legend_row, column=2).value = "Integração raster-vetor"

# Save workbook
output_path = r"c:\Users\LINES-COLAB\AppData\Roaming\QGIS\QGIS3\profiles\mtl_msi\python\plugins\MTL_Tools\metodos_utilitarios.xlsx"
wb.save(output_path)
print(f"XLSX gerado com sucesso: {output_path}")
print(f"Total de métodos: {(row - 2)}")
